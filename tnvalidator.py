#!/usr/bin/env python3

#Changelog

#Release R4 - June 11, 2026
# - Cleaned up output to declutter, consolidate reporting of passed tests
# - Fixed display/reporting bug in test for scouting orders assigned to empty units
# - Changed parser to read cached computed values for cells with formulae rather than formula text string
# - New Tests:
#   - Check for multiple units with same GT assigned to herding activity
#   - Check for non-uppercase grid designation in GOTO movement Hex targets
#   - Check for non-numeric characters/strings in Quantity Field for Transfers
# - Revised Tests:
#   - Updated check for valid units assigned activities to check whether activities assigned to newly created units are
#       warrior activities (Defence, Suppression, Security, Scouting). Updated to throw error for non-warrior activities,
#       do not flag warrior activities.

#Release R3.1 - April 22, 2026
# - Fix bug in parsing scouting orders for TN3.1 spreadsheets

#Release R3 - April 11, 2026
# - Added TN3.1 Support
# - New Tests:
#   - Check for Activities with zero/blank workers assigned (including slaves/specialists)
#   - Check for activities assigned to units that were empty at start of turn (partial hack, need to build post-transfer checking)
#   - Check for scouting missions assigned to units that were empty at start of turn (partial hack, need to build post-transfer checking)

#Release R2 - November 1, 2025
# - Fixed data references for revisions to 904-05+ auto orders sheet

#Release 1 - October 16, 2025
# - Made windows resizable
# - Added larger font option
# - Licensed under GPL v3
# *** Known Issues ***
# - Does not load/parse .xls files (please request this if it would be helpful to you)

#Alpha 7 - October 7, 2025
# - Corrected valid activity checking for converted/newly created units. Warns if Unit is in Valid Units tab but
#       not Clan tab, i.e., did not exist at beginning of turn (may or may not be error). Error if in neither.
# - Added color coding to report:
#   - Category title corresponds to Err level. Black = Clean, Orange = Warn, Red = Error.
#   - If ErrorType is clean, title is black and "No errors" message is in blue for easy scanning.
#   - If ErrorType is hit, title is red for errors/orange for warnings. Specific errors black for easy reading.
# - Changed to single-window interface
# - Added some additional error handling

#Alpha 6 - September 30, 2025
# - Corrected Skill and Research checks to only consider Tribes valid units
# - Added checking for excess (>3) skill attempts by a tribe
# - Added checking for duplicate attempts of the same skill by a tribe
# - Added checking for duplicate skill priorities by a tribe

#Alpha 5 - September 29, 2025
# - Reorganized report window
# - Fixed error when no report is selected
# - Fixed column parsing stopping on blank lines

import tkinter as tk
from tkinter import filedialog as fd
#from tkinter import simpledialog as sd
from tkinter.messagebox import showerror
from tkinter import ttk
import openpyxl
#import xlrd
import pathlib

## Data parsing functions.
# Parse arbitrary number of columns.

def parseCols(activeSheet, startCol, numCols):
    sheetData=[]
    for i in range(2, activeSheet.max_row + 1):
        if numCols == 1:
            if activeSheet.cell(i,startCol).value is None:
                pass
            else:
                sheetData.append(activeSheet.cell(i,startCol).value)
        else:
            colData = []
            rowBlank = True
            for j in range(numCols):
                if activeSheet.cell(i,startCol+j).value is not None:
                    rowBlank = False
                colData.append(activeSheet.cell(i,startCol+j).value)
            if rowBlank is False:
                sheetData.append(colData)
    return sheetData

def parseColsUpper(activeSheet, startCol, numCols):
    sheetData=[]
    for i in range(2, activeSheet.max_row + 1):
        if numCols == 1:
            if activeSheet.cell(i,startCol).value is None:
                pass
            else:
                sheetData.append(str(activeSheet.cell(i,startCol).value).upper())
        else:
            colData = []
            rowBlank = True
            for j in range(numCols):
                if activeSheet.cell(i,startCol+j).value is not None:
                    rowBlank = False
                colData.append(str(activeSheet.cell(i,startCol+j).value).upper())
            if rowBlank is False:
                sheetData.append(colData)
    return sheetData

## Compares two lists/columns, case insensitive
def checkValidList(testList, validList):
    listErrors = []
    for i in range(len(testList)):
        match = None
        for vi in validList:
            if str(testList[i]).upper() == str(vi).upper():
                match = testList[i]
               
        if match is None:
            #Add 2 to correct human-readable row number for zero indexing and omitting header row 
            errorData = (i+2,testList[i])
            listErrors.append(errorData)

    return listErrors

## File access function.
#read xlsx file and pull sheets we want read into global variables, close file
def processOrdersXLSX(path):
    
    try:
        orders = openpyxl.load_workbook(path, data_only=True)
        #orders = openpyxl.load_workbook(path)
    except:
        showerror("File Input Error", "Could not open file, may be open or in use.")
        return None

    global clanSheet
    global activitySheet
    global moveSheet
    global scoutSheet
    global skillSheet
    global researchSheet
    global transferSheet
    global vgSheet
    global vuSheet
    global vaSheet
    
    #pull all sheets we are interested in
    try:
        clanSheet = orders["Clan"]
        activitySheet = orders["Tribes_Activities"]
        moveSheet = orders["Tribe_Movement"]
        scoutSheet = orders["Scout_Movement"]
        skillSheet = orders["Skill_Attempts"]
        researchSheet = orders["Research_Attempts"]
        transferSheet = orders["Transfers"]
        vgSheet = orders["Valid Goods"]
        vuSheet = orders["Valid Units"]
        vaSheet = orders["Valid Activity"]
        sheetVersion = str(orders["Instructions"].cell(1,2).value) 
    except:
        showerror("File Input Error", "Selected file is not a valid TN Order sheet.")
        return None

    if sheetVersion == "1.12":
        gameVersion = "TN3.1"
    elif sheetVersion == "1.13":
        gameVersion = "TN3"
    else:
        showerror("File Input Error", "Unsupported Game Version")
        gameVersion = None

    orders.close()
 
    return gameVersion

#main loop called when an order sheet is selected
def select_file():

    filetypes = [('TN3 XLSX Turnsheet', '*.xlsx')]
    
    #For when XLS support is added
    #filetypes = (
    #   ('TN3 XLSX Turnsheet', '*.xlsx'),
    #    ('TN3 XLS Turnsheet', '*.xls')
    #)

    ordersPath = fd.askopenfilename(
        title='Select TN3 Order File',
        filetypes=filetypes)
    
    #If no file is selected, bug out
    if ordersPath == "":
        return None
    
    path = pathlib.Path(ordersPath)

    #destroy any existing frame if this is a second run
    for widget in root.winfo_children():
        if isinstance(widget, tk.Frame):
            widget.destroy()
            root.geometry('600x200')
            root.update()
    
    #process file, load data into globals, return game version
    if path.suffix == ".xlsx":
        gameVersion = processOrdersXLSX(path)

    #processOrdersXLSX will return None if the file cannot be opened or was invalid. Bail out.
    if gameVersion == None:
        return None

    #build results window

    results = tk.Frame(root)
    results.pack(fill = "both", expand = True)

    style = ttk.Style()
    style.configure("Treeview", font=("Arial", fontsize), rowheight=int(fontsize*2.2))
    
    treeview = ttk.Treeview(results, show="tree")
    treeview.column("#0", width=550, stretch=True)
    treescroll=ttk.Scrollbar(results, orient="vertical", command=treeview.yview)
    treeview.configure(yscrollcommand=treescroll.set)
    treeview.tag_configure("error", foreground="red")
    treeview.tag_configure("warning", foreground="orange")
    treeview.tag_configure("pass", foreground="blue")
    treeview.tag_configure("old", font=("Arial", 12))
    root.geometry('600x750')
    titleMessage= "Validating File: " + path.name
    tk.Label(results, text=titleMessage, font=("Arial", 12, "bold")).pack()

    #Create lists based on game
    
    if gameVersion == "TN3":
        clanUnitList = parseCols(clanSheet,2,1)
        activityUnitList = parseCols(activitySheet,2,1)
        movementUnitList = parseCols(moveSheet,2,1)
        scoutUnitList = parseCols(scoutSheet,2,1)
        skillUnitList = parseCols(skillSheet,1,1)
        researchUnitList = parseCols(researchSheet,1,1)
        validUnits = parseCols(vuSheet,1,1)
        skillAttempts = parseCols(skillSheet,1,3)
        clanActivities = parseCols(activitySheet,2,7)
        validActivities = parseCols(vaSheet,1,3)
        clanTransfers = parseCols(transferSheet,1,5)
        validGoods = parseCols(vgSheet,1,1)
        clanUnits = parseCols(clanSheet,2,14)
        clanMovement = parseCols(moveSheet,2,5)
        clanScouting = parseCols(scoutSheet,2,5)

    if gameVersion == "TN3.1":
        clanUnitList = parseCols(clanSheet,1,1)
        activityUnitList = parseCols(activitySheet,1,1)
        movementUnitList = parseCols(moveSheet,1,1)
        scoutUnitList = parseCols(scoutSheet,1,1)
        skillUnitList = parseCols(skillSheet,1,1)
        researchUnitList = parseCols(researchSheet,1,1)
        validUnits = parseCols(vuSheet,1,1)
        skillAttempts = parseCols(skillSheet,1,3)
        clanActivities = parseCols(activitySheet,1,7)
        validActivities = parseColsUpper(vaSheet,1,3)
        clanTransfers = parseCols(transferSheet,1,5)
        validGoods = parseCols(vgSheet,1,1)
        clanUnits = parseCols(clanSheet,1,14)
        clanMovement = parseCols(moveSheet,1,5)
        clanScouting = parseCols(scoutSheet,1,5)
        
    #Separate GM and Clan units from Valid Units
    clanNumber = str(clanUnitList[0][1:4])
    validClanUnits = []
    validGMUnits = []
    for i in range(len(validUnits)):
        if validUnits[i][1:4] == clanNumber:
            validClanUnits.append(validUnits[i])
        else:
            validGMUnits.append(validUnits[i])
        
    tk.Label(results, text=gameVersion + " Orders for Clan " + clanNumber, font=("Arial", 12, "bold")).pack()

    treeview.pack(side="left", fill="both", expand=True)
    treescroll.pack(side="left",fill="y")

    #Select Tribe units from Valid Clan Units
    validClanTribes = []
    for i in range(len(validClanUnits)):
        if len(validClanUnits[i]) == 4:
            validClanTribes.append(validClanUnits[i])
    
    #Display Valid Units
    vuRoot = treeview.insert("",0,text="Valid Units")
    cuRoot = treeview.insert(vuRoot,tk.END,text="Valid Clan Units")
    for i in range(len(validClanUnits)):
        treeview.insert(cuRoot,tk.END,text=str(validClanUnits[i]))
    vtRoot = treeview.insert(vuRoot,tk.END,text="Valid Clan Tribes" )
    for i in range(len(validClanTribes)):
        treeview.insert(vtRoot,tk.END,text=str(validClanTribes[i]))
    guRoot = treeview.insert(vuRoot,tk.END,text="Valid GM Units")
    for i in range(len(validGMUnits)):
        treeview.insert(guRoot,tk.END,text=str(validGMUnits[i]))

    #create organizational root
    actRoot = treeview.insert("",tk.END,text="Activity Orders",tags="pass")
    xfrRoot = treeview.insert("",tk.END,text="Transfer Orders",tags="pass")
    movRoot = treeview.insert("",tk.END,text="Movement and Scouting Orders",tags="pass")
    lrnRoot = treeview.insert("",tk.END,text="Skill and Research Orders",tags="pass")
    
    ### Movement and Scouting Tests
    passRoot = treeview.insert(movRoot,tk.END,text="Tests Passed",tags="pass")
    
    #Check for Invalid Units Assigned Movement Orders 
    vErrors = checkValidList(movementUnitList, validClanUnits)
    if vErrors:
        errRoot = treeview.insert(movRoot,0,text="Invalid Units Assigned Movement Orders",tags="error",open=True)
        treeview.item(movRoot, tags="error")
        for i in range(len(vErrors)):
            errorText = "Invalid Unit " + str(vErrors[i][1]) + " assigned movement order on Row " + str(vErrors[i][0])
            treeview.insert(errRoot,0,text=errorText)
    else:
        treeview.insert(passRoot,tk.END,text="No Invalid Units Assigned Movement Orders")
    
    #Check for GOTO order where Grid ID is not all uppercase
    vErrors = []
    for i in range(len(clanMovement)):
        if str(clanMovement[i][4]).upper() == "GOTO" and str(clanMovement[i][3]).isupper() is False:
            errorData = [i+2, clanMovement[i][0], clanMovement[i][3]]
            vErrors.append(errorData)
    
    if vErrors:
        errRoot = treeview.insert(movRoot,0,text="GOTO Orders With Non-Uppercase Hex Target",tags="error",open=True)
        treeview.item(movRoot, tags="error")
        for i in range(len(vErrors)):
            errorText = "Unit " + str(vErrors[i][1]) + " assigned GOTO order with non-uppercase Hex target " + str(vErrors[i][2]) + " on Row " + str(vErrors[i][0])
            treeview.insert(errRoot,0,text=errorText)
    else:
        treeview.insert(passRoot,tk.END,text="No GOTO Orders With Non-Uppercase Hex Target")
    
    #Check for Invalid Units Assigned Scouting Orders
    vErrors = checkValidList(scoutUnitList, validClanUnits)
    if vErrors:
        errRoot = treeview.insert(movRoot,0,text="Invalid Units Assigned Scouting Orders",tags="error",open=True)
        treeview.item(movRoot, tags="error")
        for i in range(len(vErrors)):
            errorText = "Invalid Unit " + str(vErrors[i][1]) + " assigned scouting order on Row " + str(vErrors[i][0])
            treeview.insert(errRoot, tk.END, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text="No Invalid Units Assigned Scouting Orders")
    
    #Check for scouting missions assigned to unit that was empty at turn start (likely absorbed or disbanded)
    vErrors = []
    emptyUnits = []
    for i in range(len(clanUnits)):
        if clanUnits[i][0] is not None:
            unitPopulation = int(clanUnits[i][2]) + int(clanUnits[i][3]) + int(clanUnits[i][3])
            if unitPopulation == 0:
                emptyUnits.append(str(clanUnits[i][0]).upper())
    assignedScouts = []
    for i in range(len(clanScouting)):
        if any(x is not None for x in clanScouting[i][1:4]):
            scoutingData = [i+2, str(clanScouting[i][0]).upper()]
            assignedScouts.append(scoutingData)  
    
    for i in range(len(assignedScouts)):
        if assignedScouts[i][1] in emptyUnits:
            vErrors.append(assignedScouts[i])
    
    if vErrors:
        errRoot = treeview.insert(movRoot, 0, text="Scouting Missions Assigned to Units That Began Turn Empty", tags="error", open=True)
        treeview.item(movRoot, tags="error")
        for i in range(len(vErrors)):
            errorText = "Unit " + str(vErrors[i][1]).lower() + " was empty at turn start and is assigned a scouting mission on row " + str(vErrors[i][0])
            treeview.insert(errRoot, tk.END, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text="No Scouting Missions Assigned to Units That Began Turn Empty")
    
    # Check for otherwise complete scouting order (i.e., scouts, horses, mission defined) with EMPTY in Movement1
    
    # Check for Scouting Missions that exceed available warriors post-transfer
    ## This needs to be finished
      
    """
    errRoot = treeview.insert(movRoot, tk.END, text="Isnufficient Warriors Scouting Errors", open=True)
    vErrors = {}

    #create dict of how many warriors are in each clan unit post transfer
    unitWarriors = {}
    for i in range(len(clanUnits)):
        if clanUnits[i][0] is not None:
            currentUnit = str(clanUnits[i][0])
            unitWarriorsCount = int(clanUnits[i][2])
            for j in range(len(clanTransfers)):
                if str(clanTransfers[j][0]).upper() == currentUnit.upper() and str(clanTransfers[j][2]).upper() == "WARRIORS":
                    unitWarriorsCount -= int(clanTransfers[j][3])
                elif str(clanTransfers[j][1]).upper() == currentUnit.upper() and str(clanTransfers[j][2]).upper() == "WARRIORS":
                    unitWarriorsCount += int(clanTransfers[j][3])
            unitWarriors[currentUnit] = unitWarriorsCount

    ## How do we do newly created units?
    # for i in validUnits, if not in unitWarriors, set Count to 0 and run transfer check? Should I do this to build unitWarriors then loop
    # through it and only eval transfers once? Or do I care?


    scoutingUnits = {}
    for i in range(len(clanScouting)):
        if clanScouting[i][1] is not None:
            currentUnit = clanScouting[i][0]    
            if currentUnit in scoutingUnits:
                scoutingUnits[currentUnit] += clanScouting[i][1]
            else:
                scoutingUnits[currentUnit] = clanScouting[i][1]
        
    for key, value in scoutingUnits.items():
        if key in unitWarriors:
            if value > unitWarriors[key]:
                vErrors[key] = value
        
    """
    ### Skill and Research Tests
    passRoot = treeview.insert(lrnRoot,tk.END,text="Tests Passed",tags="pass")
    
    #Check for Invalid Units Assigned Skill Attempts
    vErrors = checkValidList(skillUnitList, validClanTribes)
    if vErrors:
        errRoot = treeview.insert(lrnRoot,0,text="Invalid Units Assigned Skill Attempts", tags="error",open=True)
        treeview.item(lrnRoot, tags="error")
        for i in range(len(vErrors)):
            errorText = "Invalid Unit " + str(vErrors[i][1]) + " assigned skill attempt on Row " + str(vErrors[i][0])
            treeview.insert(errRoot, 0, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text="No Invalid Units Assigned Skill Attempts")

    #Add - Check for Priority 3 (Teacher) Skill attempt assigned to non-Group A skill

    #Add - Check for Priority 3 (Teacher) skill attempt assigned to skill level beyond 7
    
    #check for Tribes assigned more than three skill attempts
    skillAttemptTrack = {}
    vErrors = {}
    for i in skillUnitList:
        if i in skillAttemptTrack:
            skillAttemptTrack[i] += 1
        else:
            skillAttemptTrack[i] = 1

    for key, value in skillAttemptTrack.items():
        if value > 3:
            vErrors[key] = value

    if vErrors:
        errRoot = treeview.insert(lrnRoot,0,text="Tribes Assigned More Than Three Skill Attempts",tags="error", open=True)
        treeview.item(lrnRoot, tags="error")
        for key, value in vErrors.items():
            errorText = "Tribe " + str(key) + " assigned " + str(value) + " skill attempts"
            treeview.insert(errRoot, tk.END, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text="No Tribe Assigned More Than Three Skill Attempts")
    
    #check for duplicate Tribe/Skill attempts
    skillAttemptTrack = []
    vErrors = []
    for i in range(len(skillAttempts)):
        checkAttempt = [skillAttempts[i][0], str(skillAttempts[i][2]).upper()]
        if checkAttempt in skillAttemptTrack:
            vErrors.append(checkAttempt)
        else:
            skillAttemptTrack.append(checkAttempt)
    if vErrors:
        errRoot = treeview.insert(lrnRoot,0,text="Tribes Attempting Duplicate Skills", tags="error", open=True)
        treeview.item(lrnRoot, tags="error")
        for i in range(len(vErrors)):
            errorText = "Tribe " + str(vErrors[i][0]) + " attempting duplicate skill " + str(vErrors[i][1]) 
            treeview.insert(errRoot, tk.END, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text="No Tribe Attempting Duplicate Skills")

    #check for skill attempts with same priority for Tribe
    skillAttemptTrack = []
    vErrors = []

    for i in range(len(skillAttempts)):
        checkAttempt = [skillAttempts[i][0], skillAttempts[i][1]]
        if checkAttempt in skillAttemptTrack:
            vErrors.append(checkAttempt)
        else:
            skillAttemptTrack.append(checkAttempt)
    if vErrors:
        errRoot = treeview.insert(lrnRoot,0,text="Tribes Attempting Skills at Same Priority", tags="error", open=True)
        treeview.item(lrnRoot, tags="error")
        for i in range(len(vErrors)):
            errorText = "Tribe " + str(vErrors[i][0]) + " attempting multiple skills at priority " + str(vErrors[i][1]) 
            treeview.insert(errRoot, tk.END, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text="No Tribe Attempting Skills At Same Priority")

    #Check for Invalid Units Assigned Research Attempts
    vErrors = checkValidList(researchUnitList, validClanTribes)
    if vErrors:
        errRoot = treeview.insert(lrnRoot,0,text="Invalid Units Assigned Research Attempts", tags="error", open=True)
        treeview.item(lrnRoot, tags="error")
        for i in range(len(vErrors)):
            errorText = "Invalid Unit " + str(vErrors[i][1]) + " assigned research attempt on Row " + str(vErrors[i][0])
            treeview.insert(errRoot, tk.END, text=errorText)
            #DEBUG CODE
            #debugText = "Data type: " + str(type((vErrors[i][1])))
            #treeview.insert(errRoot, tk.END, text=debugText)
            #codes = [str(ord(char)) for char in str(vErrors[i][1])]
            #debugText = "Character Codes: " + " ".join(codes)
            #treeview.insert(errRoot, tk.END, text=debugText)
    else:
        treeview.insert(passRoot, tk.END, text="No Invalid Units Assigned Research Attempts")
    
    ### Activity Tests
    passRoot = treeview.insert(actRoot,tk.END,text="Tests Passed",tags="pass")
    
    #Check for Invalid Units Assigned Activities
  
    #Check clan tab first because new units ordinariliy should not perform activities. If unit is on valid list, give warning (converted unit, scouting orders). If not on valid list, give error.
    vErrors = checkValidList(activityUnitList, clanUnitList)
    newUnitErrors = []
    invalidUnitErrors = []
    validNewUnitActivities = ["DEFENCE", "SCOUTING", "SUPPRESSION", "SECURITY"]

    if vErrors:
        for i in range(len(vErrors)):
            if vErrors[i][1] in validClanUnits:
                adjustedRowNumber = int(vErrors[i][0])-2
                newUnitActivity = clanActivities[adjustedRowNumber][1]
                if newUnitActivity not in validNewUnitActivities:
                    errorData = [vErrors[i][0], vErrors[i][1], newUnitActivity]
                    newUnitErrors.append(errorData)
            else:
                invalidUnitErrors.append(vErrors[i])

    if newUnitErrors:
        errRoot = treeview.insert(actRoot,0,text="New Unit Assigned Invalid (Non-Warrior) Activity Orders", tags="error", open=True)
        treeview.item(actRoot, tags="error")
        for i in range(len(newUnitErrors)):
            errorText = "Newly Created Unit " + str(newUnitErrors[i][1]) + " assigned invalid activity " + str(newUnitErrors[i][2]) + " on Row " + str(newUnitErrors[i][0])
            treeview.insert(errRoot, tk.END, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text="No New Units Assigned Invalid Activity Orders")

    if invalidUnitErrors:
        errRoot = treeview.insert(actRoot,0,text="Invalid Unit Assigned Activity Orders", tags="error", open=True)
        treeview.item(actRoot, tags="error")
        for i in range(len(invalidUnitErrors)):
            errorText = "Invalid Unit " + str(invalidUnitErrors[i][1]) + " assigned activity on Row " + str(invalidUnitErrors[i][0])
            treeview.insert(errRoot, tk.END, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text="No Invalid Units Assigned Activity Orders")

    #Check for invalid Activities
    vErrors = []
    for i in range(len(clanActivities)):
        
        casedActivity = [str(clanActivities[i][1]).upper(), str(clanActivities[i][2]).upper(), str(clanActivities[i][3]).upper()]
        
        if clanActivities[i][2] is None or clanActivities[i][3] is None:
            if clanActivities[i][2] is None:
                item = "[Item Blank]"
            else:
                item = clanActivities[i][2]

            if clanActivities[i][3] is None:
                distinction = "[Distinction Blank]"
            else:
                distinction = clanActivities[i][3]

            errorData = [i+2, clanActivities[i][1], item, distinction]
            vErrors.append(errorData)
              
        elif casedActivity not in validActivities:     
            errorData = [i+2, clanActivities[i][1], clanActivities[i][2], clanActivities[i][3]]
            vErrors.append(errorData)

    if vErrors:
        errRoot = treeview.insert(actRoot,0, text="Invalid Activity/Item/Distinction Groupings", tags="error", open=True)
        treeview.item(actRoot, tags="error")
        for i in range(len(vErrors)):
            errorText = "Invalid Activity/Item/Distinction on Row " + str(vErrors[i][0]) + ": " + str(vErrors[i][1]) + " / " + str(vErrors[i][2]) + " / " + str(vErrors[i][3])
            treeview.insert(errRoot, tk.END, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text="No Invalid Activity/Item/Distinction Groupings")
    
    #check for Activity Discontinuity
    actAssignedUnits = []
    vErrors = []
    prevActUnit = None

    for i in range(len(activityUnitList)):
        curUnit = activityUnitList[i]
        if curUnit not in actAssignedUnits:
            prevActUnit=curUnit
            actAssignedUnits.append(curUnit)
        
        else:
            if curUnit == prevActUnit:
                actAssignedUnits.append(curUnit)
            else:
                prevActUnit=curUnit
                actAssignedUnits.append(curUnit)
                errorData = (i+2, curUnit)
                vErrors.append(errorData)
    
    if vErrors:
        errRoot = treeview.insert(actRoot,0,text="Fragmented Activity Orders", tags="error", open=True)
        treeview.item(actRoot, tags="error")
        for i in range(len(vErrors)):
            errorText = "Unit " + str(vErrors[i][1]) + " assigned non-contiguous activity order on Row " + str(vErrors[i][0])
            treeview.insert(errRoot, tk.END, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text="No Fragmented Activity Orders")

    #check for Activities assigned no workers
    vErrors = []
    for i in range(len(clanActivities)):
        try:
            peopleCount = int(clanActivities[i][4])
        except:
            peopleCount = 0
        try:
            peopleCount += int(clanActivities[i][5])
        except:
            peopleCount += 0
        try:
            peopleCount += int(clanActivities[i][6])
        except:
            peopleCount += 0
    
        if peopleCount <= 0:
            errorData = (i+2, clanActivities[i][0], clanActivities[i][1])
            vErrors.append(errorData)
    
    if vErrors:
        errRoot = treeview.insert(actRoot, 0, text="Activity Orders With Fewer than 1 Worker Assigned", tags="error", open=True)
        treeview.item(actRoot, tags="error")
        for i in range(len(vErrors)):
            errorText = "Fewer than 1 Worker Assigned to Unit " + str(vErrors[i][1]).lower() + " Activity " + str(vErrors[i][2]) + " on Row " + str(vErrors[i][0])
            treeview.insert(errRoot, tk.END, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text = "No Activity Orders With Fewer than 1 Worker Assigned")

    #check for multiple units in same GT assigned to herding
    vErrors = []
    herdingUnits = []
    herdedUnits = {}
    unitGTs = {}

    for i in range(len(clanUnits)):
        unit = str(clanUnits[i][0]).upper()
        gt = str(clanUnits[i][1]).upper()
        unitGTs[unit] = gt
    
    for i in range(len(clanActivities)):
        if str(clanActivities[i][1]).upper() == "HERDING":
            herdingUnits.append(str(clanActivities[i][0]).upper())

    for i in range(len(herdingUnits)): #walk through all units assigned herding activities
        herdingUnit = herdingUnits[i] 
        if herdingUnit in unitGTs: #doing this to avoid potential errors
            herdedUnit = str(unitGTs[herdingUnit]).upper() #set the "herded units" for the herding activity at issue to the GT of the herding unit

            if herdedUnit in herdedUnits: # if the herded unit is already a key in the herded units dict
                existingHerder = herdedUnits[herdedUnit] # identify existing herder
                errorData = [herdingUnit.lower(), herdedUnit.lower(), existingHerder.lower()] # add herding unit, herded unit, and existing herder to errors
                vErrors.append(errorData)
            else:
                herdedUnits[herdedUnit] = herdingUnit # otherwise add herded unit as key in herded unit with value of herder

    if vErrors:
        errRoot = treeview.insert(actRoot, 0, text="Duplicative Herding Activities For Same Goods Tribe", tags="error", open=True)
        treeview.item(actRoot, tags="error")

        for i in range(len(vErrors)):
            errorText = "Unit " + str(vErrors[i][0]) + " assigned to herd animals of GT " + str(vErrors[i][1]) + " that Unit " + str(vErrors[i][2]) + " is already herding."
            treeview.insert(errRoot, tk.END, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text="No Duplicative Herding Activities For Same Goods Tribe")
    
    #check for Activities assigned to Unit that was Empty at start of turn (likely absorbed our disbanded but persisting)
    vErrors = []
    emptyUnits = []
    for i in range(len(clanUnits)):
        if clanUnits[i][0] is not None:
            unitPopulation = int(clanUnits[i][2]) + int(clanUnits[i][3]) + int(clanUnits[i][3])
            if unitPopulation == 0:
                emptyUnits.append(str(clanUnits[i][0]).upper())

    for i in range(len(clanActivities)):
        if str(clanActivities[i][0]).upper() in emptyUnits:
            errorData = (i+2, clanActivities[i][0], clanActivities[i][1])
            vErrors.append(errorData)
    
    if vErrors:
        errRoot = treeview.insert(actRoot, 0, text="Activities Assigned to Units That Began Turn Empty", tags="error", open=True)
        treeview.item(actRoot, tags="error")

        for i in range(len(vErrors)):
            errorText = "Unit " + str(vErrors[i][1]) + " was empty at turn start and is assigned activity " + str(vErrors[i][2]) + " on row " + str(vErrors[i][0])
            treeview.insert(errRoot, tk.END, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text="No Activities Assigned to Units That Began Turn Empty")

    ##Transfer Orders
    passRoot = treeview.insert(xfrRoot,tk.END,text="Tests Passed",tags="pass")
    
    #check for at least one Clan unit in each transfer
    vErrors = []
    for i in range (len(clanTransfers)):
        match = None
        for vi in validClanUnits:
            if str(clanTransfers[i][0]).upper() == vi.upper():
                match = clanTransfers[i][0]
            if str(clanTransfers[i][1]).upper() == vi.upper():
                match = clanTransfers[i][1]
        
        if match is None:
            #add 2 for omitted title row and zero index conversion
            vErrors.append(i+2)
             
    if vErrors:
        errRoot = treeview.insert(xfrRoot,0,text="Transfer Orders Without Clan Unit", tags="error", open=True)
        treeview.item(xfrRoot, tags="error")

        for i in range(len(vErrors)):
            errorText = "Transfer order on Row " + str(vErrors[i]) + " has no valid Clan unit"
            treeview.insert(errRoot, tk.END, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text="No Transfer Orders Without Clan Unit")

    #check for transfers from non-Clan/GM Units (not a valid transfer order, error)

    vErrors = []
    for i in range (len(clanTransfers)):
        match = None
        for vi in validUnits:
            if str(clanTransfers[i][0]).upper() == vi.upper():
                match = clanTransfers[i][0]

        if match is None:
            errorData = (i+2, clanTransfers[i][0], clanTransfers[i][1])
            vErrors.append(errorData)

    if vErrors:
        errRoot = treeview.insert(xfrRoot,0,text="Transfers from External (Non-Clan/GM) Units", tags="error", open=True)
        treeview.item(xfrRoot, tags="error")

        for i in range(len(vErrors)):
            errorText = "Transfer From Non-Clan/GM Unit " + str(vErrors[i][1]) + " to Unit " + str(vErrors[i][2]) + " on Row " + str(vErrors[i][0])
            treeview.insert(errRoot, tk.END, text=errorText)
    else:   
        treeview.insert(passRoot, tk.END, text="No Transfers from External (Non-Clan/GM) Units")

    #check for transfers to non-Clan Units (valid but worth reviewing for mistakes, warning)
    vErrors = []
    for i in range (len(clanTransfers)):
        match = None
        for vi in validUnits:
            if str(clanTransfers[i][1]).upper() == vi.upper():
                match = clanTransfers[i][1]

        if match is None:
            errorData = (i+2, clanTransfers[i][0], clanTransfers[i][1])
            vErrors.append(errorData)

    if vErrors:
        errRoot = treeview.insert(xfrRoot,0,text="Transfers to External (Non-Clan/GM) Units [Warning Only]", tags="warning", open=True)

        if treeview.item(xfrRoot, option="tags") != ("error",):
            treeview.item(xfrRoot, tags="warning")

        for i in range(len(vErrors)):
            errorText = "Transfer To Non-Clan/GM Unit " + str(vErrors[i][2]) + " from Unit " + str(vErrors[i][1]) + " on Row " + str(vErrors[i][0])
            treeview.insert(errRoot, tk.END, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text="No Transfers to External (Non-Clan/GM) Units")    

    #check for invalid goods in transfers
    vErrors = []
    for i in range (len(clanTransfers)):
        match = None
        for vi in validGoods:
            if str(clanTransfers[i][2]).upper() == vi.upper():
                match = clanTransfers[i][2]

        if match is None:
            errorData = (i+2, clanTransfers[i][2])
            vErrors.append(errorData)

    if vErrors:
        errRoot = treeview.insert(xfrRoot,0,text="Invalid Goods in Transfer Orders", tags="error", open=True)
        treeview.item(xfrRoot, tags="error")

        for i in range(len(vErrors)):
            errorText = "Invalid Good " + str(vErrors[i][1]) + " on Row " + str(vErrors[i][0])
            treeview.insert(errRoot, tk.END, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text="No Invalid Goods in Transfer Orders")

    # Check for transfer with blank "From" Unit (Pend this item for if parsecols is later changed not to flag an empty row where acting unit is missing)
    # Check for transfer with blank "To" Unit
    # Check for transfer with blank Item
    # Check for transfer with blank Quantity
    
    # Check for transfer with invalid (non-numeric) quantity
    vErrors = []

    for i in range (len(clanTransfers)):
        if not isinstance(clanTransfers[i][3], int):
            errorData = [i+2, clanTransfers[i][3]]
            vErrors.append(errorData)

        elif str(clanTransfers[i][3]).isdigit() is False:
            errorData = [i+2, clanTransfers[i][3]]
            vErrors.append(errorData)
    
    if vErrors:
        errRoot = treeview.insert(xfrRoot,0,text="Transfer Orders with Non-Numeric Quantity", tags="error", open=True)
        treeview.item(xfrRoot, tags="error")

        for i in range(len(vErrors)):
            errorText = "Invalid (Non-Numeric) Characters or String in Quantity Field for Transfer on Row " + str(vErrors[i][0]) + ": " + str(vErrors[i][1])
            treeview.insert(errRoot, tk.END, text=errorText)
    else:
        treeview.insert(passRoot, tk.END, text="No Transfer Orders with Non-Numeric Quantity")

    # Check for transfer with blank Transfer_Timing

fontsize = 10

#fontchange
def font_resize():
    global fontsize
    fontsize = checkVar.get()

#create main window

root = tk.Tk()
root.title('TNValidator')
root.resizable(True, True)
root.geometry('640x200')

tk.Label(text="TN Order Validator by Clan 293", font=("Arial", 12, "bold")).pack()
tk.Label(text="Release R4, 2026-06-11", font=("Arial", 10, "bold")).pack()

#Font doodling
checkVar = tk.IntVar()
fontcheck = tk.Checkbutton(root, text="Check for Larger Font", variable = checkVar, onvalue = 12, offvalue = 10, command=font_resize)

tk.Label(text="Please select your TN Auto Order Sheet to Validate").pack()

# open button
open_button = tk.Button(
    root,
    text='Select File',
    command=select_file
)

open_button.pack()

tk.Label(text="Add New Units to Column A of Valid Units Sheet to Reduce False Positives").pack()
fontcheck.pack()
fontcheck.deselect()

root.mainloop()